from __future__ import annotations

from io import BytesIO
from openpyxl import load_workbook


def extract_text_from_xlsx_bytes(xlsx_bytes: bytes, max_rows_per_sheet: int = 2000) -> str:
    """
    Extract text from XLSX.
    Strategy: iterate sheets, dump non-empty cell values row-wise.
    max_rows_per_sheet prevents huge files from blowing up.
    """
    wb = load_workbook(filename=BytesIO(xlsx_bytes), data_only=True, read_only=True)
    parts: list[str] = []

    for ws in wb.worksheets:
        parts.append(f"--- SHEET: {ws.title} ---")
        row_count = 0

        for row in ws.iter_rows(values_only=True):
            if row is None:
                continue
            # Build a compact row string
            cells = []
            for v in row:
                if v is None:
                    continue
                s = str(v).strip()
                if s:
                    cells.append(s)

            if cells:
                parts.append(" | ".join(cells))
                row_count += 1

            if row_count >= max_rows_per_sheet:
                parts.append(f"[Truncated after {max_rows_per_sheet} rows]")
                break

        parts.append("")  # spacer

    return "\n".join(parts).strip()
